import re
import os
import openpyxl as op


def get_dict():
        #  打开excel
        wb = op.load_workbook('汇总.xlsx')
        ws = wb['Sheet1']
        all_list = []
        for i in range(2,ws.max_row+1):
            dict_host = {}
            dict_host['sys_name'] = (ws.cell(row=i,column=1)).value
            dict_host['ipadd'] = (ws.cell(row=i,column=2)).value
            dict_host['area'] = (ws.cell(row=i,column=3)).value
            dict_host['principal'] = (ws.cell(row=i,column=4)).value
            dict_host['tel'] = (ws.cell(row=i,column=5)).value
            dict_host['role'] = (ws.cell(row=i,column=6)).value

            all_list.append(dict_host)
        # for i in range(0, len(all_list)):
        #     print(all_list[i])

        wb2 = op.load_workbook('workloadconfig.xlsx')
        ws2 = wb2['workloadconfig']
        for i in range(2, ws2.max_row + 1):
            ip = (ws2.cell(row=i,column=5)).value
            for x in range(len(all_list)):
                if all_list[x].get('ipadd') == ip:
                    ws2.cell(row=i, column=6, value=str(all_list[x].get('role')))
                    ws2.cell(row=i, column=8, value=str(all_list[x].get('area')))
                    ws2.cell(row=i, column=9, value=str(all_list[x].get('sys_name')))
        wb2.save('workloadconfig1.csv')


if __name__ == '__main__':
    get_dict()